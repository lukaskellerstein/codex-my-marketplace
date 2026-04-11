#!/usr/bin/env python3
"""
XLSX Preview Script

Generates a terminal-friendly preview of an Excel spreadsheet for QA.
Shows sheet names, dimensions, first N rows, formula cells, and error cells.

Usage:
    python preview.py <excel_file> [--rows N] [--formulas] [--errors-only]
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


EXCEL_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def preview_sheet(ws, max_rows=10, show_formulas=False):
    """Generate preview data for a single worksheet."""
    info = {
        "name": ws.title,
        "dimensions": ws.dimensions or "empty",
        "rows": ws.max_row or 0,
        "cols": ws.max_column or 0,
    }

    # Collect rows
    rows_data = []
    formula_cells = []
    error_cells = []

    for row_idx, row in enumerate(ws.iter_rows(max_row=min(max_rows, ws.max_row or 0)), start=1):
        row_values = []
        for cell in row:
            val = cell.value
            if val is not None and isinstance(val, str):
                if val.startswith("="):
                    formula_cells.append(f"{cell.coordinate}: {val}")
                for err in EXCEL_ERRORS:
                    if err in str(val):
                        error_cells.append(f"{cell.coordinate}: {val}")
                        break
            row_values.append(str(val) if val is not None else "")
        rows_data.append(row_values)

    # Scan remaining rows for formulas/errors if beyond preview range
    if ws.max_row and ws.max_row > max_rows:
        for row in ws.iter_rows(min_row=max_rows + 1):
            for cell in row:
                val = cell.value
                if val is not None and isinstance(val, str):
                    if val.startswith("="):
                        formula_cells.append(f"{cell.coordinate}: {val}")
                    for err in EXCEL_ERRORS:
                        if err in str(val):
                            error_cells.append(f"{cell.coordinate}: {val}")
                            break

    info["preview_rows"] = rows_data
    info["formula_count"] = len(formula_cells)
    info["formulas"] = formula_cells[:20] if show_formulas else []
    info["error_count"] = len(error_cells)
    info["errors"] = error_cells[:20]

    return info


def print_table(rows, col_widths=None):
    """Print rows as a simple ASCII table."""
    if not rows:
        return

    if col_widths is None:
        num_cols = max(len(r) for r in rows)
        col_widths = [0] * num_cols
        for row in rows:
            for i, val in enumerate(row):
                if i < num_cols:
                    col_widths[i] = max(col_widths[i], min(len(val), 30))

    for row in rows:
        cells = []
        for i, val in enumerate(row):
            width = col_widths[i] if i < len(col_widths) else 10
            truncated = val[:width].ljust(width) if len(val) <= width else val[:width - 2] + ".."
            cells.append(truncated)
        print(" | ".join(cells))


def preview(filename, max_rows=10, show_formulas=False, errors_only=False, as_json=False):
    """Generate and display preview of an XLSX file."""
    path = Path(filename)
    if not path.exists():
        print(f"Error: {filename} does not exist", file=sys.stderr)
        sys.exit(1)

    # Load with formulas (not data_only) to see formula strings
    wb = load_workbook(filename, data_only=False)

    all_sheets = []
    total_formulas = 0
    total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = preview_sheet(ws, max_rows, show_formulas)
        all_sheets.append(info)
        total_formulas += info["formula_count"]
        total_errors += info["error_count"]

    wb.close()

    if as_json:
        result = {
            "file": str(path),
            "sheet_count": len(all_sheets),
            "total_formulas": total_formulas,
            "total_errors": total_errors,
            "sheets": all_sheets,
        }
        print(json.dumps(result, indent=2))
        return

    # Print human-readable preview
    print(f"File: {path.name}")
    print(f"Sheets: {len(all_sheets)} | Formulas: {total_formulas} | Errors: {total_errors}")
    print("=" * 60)

    for info in all_sheets:
        print(f"\n--- {info['name']} ({info['rows']}x{info['cols']}, {info['dimensions']}) ---")

        if not errors_only and info["preview_rows"]:
            print_table(info["preview_rows"])
            if info["rows"] > max_rows:
                print(f"  ... {info['rows'] - max_rows} more rows")

        if info["error_count"] > 0:
            print(f"\n  ERRORS ({info['error_count']}):")
            for err in info["errors"]:
                print(f"    {err}")

        if show_formulas and info["formulas"]:
            print(f"\n  FORMULAS ({info['formula_count']}):")
            for f in info["formulas"]:
                print(f"    {f}")
            if info["formula_count"] > 20:
                print(f"    ... {info['formula_count'] - 20} more")

    if total_errors > 0:
        print(f"\nWARNING: {total_errors} error(s) found across all sheets!")
    else:
        print(f"\nAll clean — no errors found.")


def main():
    parser = argparse.ArgumentParser(description="Preview an XLSX file for QA")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--rows", type=int, default=10, help="Number of rows to preview per sheet (default: 10)")
    parser.add_argument("--formulas", action="store_true", help="Show formula cells")
    parser.add_argument("--errors-only", action="store_true", help="Only show error cells")
    parser.add_argument("--json", action="store_true", help="Output as JSON")

    args = parser.parse_args()
    preview(args.file, args.rows, args.formulas, args.errors_only, args.json)


if __name__ == "__main__":
    main()
